from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from copy import deepcopy
import pandas as pd
import numpy as np
import P0_XmlToCsv as P0

# Update CSV
P0.UpdateCSV()

# Call CSV
New_data_from_mendeley = pd.read_csv("new.csv")

# Append values to Main Excel
filename = "Reading_Ratings.xlsx"

workbook = load_workbook(filename) 
  
# Get the first sheet
sheet = workbook.worksheets[0] 

# Grab all tables on the sheet
tables = []

for table in sheet.tables.values():
    tables.append(table)

# Grab loc and style of table of interest (TOI)
toi = tables[0]
style = deepcopy(toi.tableStyleInfo)

# Set font style Variable
fontStyle = Font(size = "20")

# Add CSV data to excel file
# BUG FIX 3: Reset 'max_rows' from "max rows on sheet" to "max rows on sheet with data" 
Total_Rows = sheet.max_row
Index_for_cleaning_up = int(toi.ref.split(':')[1].strip('K'))
sheet.delete_rows(Index_for_cleaning_up,10000)
cleaned_up_rows = sheet.max_row

# Create a list to store the values 
titles = [] 
  
# Iterate through columns 
for column in sheet.iter_cols(): 
    # Get the value of the first cell in the 
    # column (the cell with the column name) 
    column_name = column[0].value 
    # Check if the column is the "Name" column 
    if column_name == "Title": 
        # Iterate over the cells in the column 
        for cell in column: 
            # Add the value of the cell to the list 
            titles.append(cell.value) 

New_additions = 0
for index, title in enumerate(New_data_from_mendeley['Title']):
    if title not in titles:
        New_additions += 1
        sheet.append(New_data_from_mendeley.iloc[index].values.tolist())    #^ File Edit

# Adjust table size to match csv | (+1) to value to prevent overwriting of Header Row and maintain consistency
Rows_of_Dataset = len(New_data_from_mendeley)
toi.ref = f"A1:K{Rows_of_Dataset + 1}"                                      #^ File Edit

# BUG FIX 1: Integers in table number 10 to cause problems when processing in Microsoft Excel, therefore, table name was adapted to avoid Integers.
toi.name = "Table"                                                          #^ File Edit

# BUG FIX 2: Integers in table number 10 to cause problems when processing in Microsoft Excel, therefore, table name was adapted to avoid Integers.
for cells in sheet[f"A1:K{Rows_of_Dataset + 1}"]:                            #^ File Edit
    for cell in cells:
        cell.font = fontStyle     

toi.tableStyleInfo = style                                              #^ File Edit

workbook.save(filename)

print(f"rows removed(rows without data) = {Total_Rows-cleaned_up_rows}")
print(f"New additions = {New_additions}")